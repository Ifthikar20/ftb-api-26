import csv
import io
import logging
from io import BytesIO

from apps.leads.models import Lead, LeadNote
from core.exceptions import ResourceNotFound
from core.logging.audit_logger import audit_log

logger = logging.getLogger("apps")


class LeadService:
    @staticmethod
    def get_leads(*, website_id: str, status: str = None, min_score: int = None):
        qs = Lead.objects.filter(website_id=website_id).order_by("-score")
        if status:
            qs = qs.filter(status=status)
        if min_score is not None:
            qs = qs.filter(score__gte=min_score)
        return qs

    @staticmethod
    def get_lead(*, website_id: str, lead_id: str) -> Lead:
        try:
            return Lead.objects.get(id=lead_id, website_id=website_id)
        except Lead.DoesNotExist:
            raise ResourceNotFound("Lead not found.") from None

    @staticmethod
    def update_status(*, lead: Lead, status: str, user) -> Lead:
        lead.status = status
        lead.save(update_fields=["status", "updated_at"])
        audit_log("lead.status_updated", user=user, action="update", resource_type="lead", resource_id=str(lead.id), metadata={"status": status, "website_id": str(lead.website_id)})

        # Dispatch outbound webhook for downstream integrations
        try:
            from apps.websites.services.webhook_service import WebhookService
            WebhookService.dispatch(
                website=lead.website,
                event="lead.status_changed",
                payload={
                    "lead_id": str(lead.id),
                    "status": status,
                    "score": lead.score,
                    "email": lead.email,
                    "company": lead.company,
                    "website_id": str(lead.website_id),
                },
            )
        except Exception as e:
            logger.warning("Webhook dispatch failed for lead.status_changed: %s", e)

        return lead

    @staticmethod
    def add_note(*, lead: Lead, content: str, user) -> LeadNote:
        return LeadNote.objects.create(lead=lead, author=user, content=content)

    @staticmethod
    def export_csv(*, website_id: str) -> str:
        leads = Lead.objects.filter(website_id=website_id).select_related("visitor")
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["ID", "Score", "Status", "Email", "Name", "Company", "Source", "Country", "Created At"])
        for lead in leads:
            writer.writerow([
                str(lead.id), lead.score, lead.status, lead.email, lead.name,
                lead.company, lead.source,
                lead.visitor.geo_country if lead.visitor else "",
                lead.created_at.isoformat(),
            ])
        audit_log("lead.exported_csv", action="export", resource_type="lead", metadata={"website_id": website_id, "count": leads.count()})
        return output.getvalue()

    @staticmethod
    def export_xlsx(*, website_id: str) -> bytes:
        """Export leads as an Excel (.xlsx) workbook. Returns raw bytes."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise RuntimeError(
                "openpyxl is not installed. Run: pip install openpyxl"
            ) from None

        leads = Lead.objects.filter(website_id=website_id).select_related("visitor")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads"

        headers = ["ID", "Score", "Status", "Email", "Name", "Company", "Source", "Country", "Created At"]
        header_font = Font(bold=True)
        header_fill = PatternFill(fill_type="solid", fgColor="4472C4")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, lead in enumerate(leads, 2):
            ws.cell(row=row_idx, column=1, value=str(lead.id))
            ws.cell(row=row_idx, column=2, value=lead.score)
            ws.cell(row=row_idx, column=3, value=lead.status)
            ws.cell(row=row_idx, column=4, value=lead.email)
            ws.cell(row=row_idx, column=5, value=lead.name)
            ws.cell(row=row_idx, column=6, value=lead.company)
            ws.cell(row=row_idx, column=7, value=lead.source)
            ws.cell(row=row_idx, column=8, value=lead.visitor.geo_country if lead.visitor else "")
            ws.cell(row=row_idx, column=9, value=lead.created_at.isoformat())

        audit_log("lead.exported_xlsx", action="export", resource_type="lead", metadata={"website_id": website_id, "count": leads.count()})

        output = BytesIO()
        wb.save(output)
        return output.getvalue()
